# -*- coding: utf-8 -*-

############################################################

#    A script to read the values of Ripoll-Sanchez et al 2023

############################################################


from cect.ConnectomeReader import ConnectionInfo
from cect.ConnectomeReader import analyse_connections

from cect.ConnectomeDataset import ConnectomeDataset

from cect.Cells import EXTRASYNAPTIC_SYN_TYPE
from cect.Cells import PEPTIDERGIC_SYN_CLASS

from cect import print_

import os
import csv


def standardise_cell_name(cell):
    for pf in ["DA", "DB", "AS", "VA", "VB", "VC", "VD", "DD"]:
        cell = cell.replace("%s0" % pf, pf)
    return cell


class RipollSanchezDataReader(ConnectomeDataset):
    verbose = False

    def __init__(self, model):
        spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"

        self.filename = "%s01022024_neuropeptide_connectome_%s.csv" % (
            spreadsheet_location,
            model,
        )

        ConnectomeDataset.__init__(self)

        cells, neuron_conns = self.read_data()
        for conn in neuron_conns:
            self.add_connection_info(conn)

    def read_data(self):
        cells = []
        conns = []

        with open(self.filename) as csvfile:
            data = list(csv.reader(csvfile))

        print_("Opened the CSV file: " + self.filename)

        for i in range(1, len(data)):
            pre_cell = standardise_cell_name(data[0][i])
            for j in range(1, len(data)):
                post_cell = standardise_cell_name(data[0][j])
                num = int(data[i][j])
                if num > 0:
                    if self.verbose:
                        print_("Conn %s->%s: %i" % (pre_cell, post_cell, num))
                    synclass = PEPTIDERGIC_SYN_CLASS
                    syntype = EXTRASYNAPTIC_SYN_TYPE

                    conns.append(
                        ConnectionInfo(pre_cell, post_cell, num, syntype, synclass)
                    )
                    if pre_cell not in cells:
                        cells.append(pre_cell)
                    if post_cell not in cells:
                        cells.append(post_cell)

        return cells, conns

    def read_muscle_data(self):
        conns = []
        neurons = []
        muscles = []

        return neurons, muscles, conns


def get_instance():
    return RipollSanchezDataReader("short_range_model")


my_instance = get_instance()

read_data = my_instance.read_data
read_muscle_data = my_instance.read_muscle_data


def load_hub_info():
    from openpyxl import load_workbook

    neuron_info_file = (
        os.path.dirname(os.path.abspath(__file__))
        + "/data/1-s2.0-S0896627323007560-mmc7.xlsx"
    )

    wb = load_workbook(neuron_info_file)
    sheet = wb.worksheets[0]
    print_("Opened the Excel file: " + neuron_info_file)

    clusters = {}

    for row in sheet.iter_rows(
        min_row=3, values_only=True
    ):  # Assuming data starts from the second row
        cell = standardise_cell_name(str(row[0]))
        # type_ = str(row[2])
        cluster = str(row[12])

        # print(f"Cell {cell} is of type {type_} and in cluster {cluster}")

        if cluster not in clusters:
            clusters[cluster] = []
        clusters[cluster].append(cell)

    return clusters


def main():
    tdr_instance = RipollSanchezDataReader("short_range_model")

    cells, neuron_conns = tdr_instance.read_data()

    neurons2muscles, muscles, muscle_conns = tdr_instance.read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)

    print_(" -- Finished analysing connections using: %s" % os.path.basename(__file__))

    print(tdr_instance.summary())

    from cect.ConnectomeView import RAW_VIEW as view
    # from cect.ConnectomeView import SOCIAL_VIEW as view
    # from cect.ConnectomeView import COOK_FIG3_VIEW as view

    cds2 = tdr_instance.get_connectome_view(view)

    print(cds2.summary())

    # fig = cds2.to_plotly_hive_plot_fig(list(view.synclass_sets.keys())[0], view)

    # fig = cds2.to_plotly_graph_fig(list(view.synclass_sets.keys())[0], view)
    fig = cds2.to_plotly_matrix_fig(
        list(view.synclass_sets.keys())[2],
        view,
        color_continuous_scale=["white", "blue", "cyan", "green", "yellow", "orange"],
    )

    import plotly.io as pio

    pio.renderers.default = "browser"
    import sys

    if "-nogui" not in sys.argv:
        fig.show()


if __name__ == "__main__":
    # main()
    load_hub_info()
