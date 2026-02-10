# -*- coding: utf-8 -*-

############################################################

#    A script to read the neurotransmitter atlas values from Wang et al. 2024.

############################################################


from cect.ConnectomeReader import analyse_connections
from cect.ConnectomeDataset import ConnectomeDataset
from cect.ConnectomeDataset import LOAD_READERS_FROM_CACHE_BY_DEFAULT
from cect.Cells import is_any_neuron
from cect.ConnectomeDataset import load_connectome_dataset_file
from cect.ConnectomeDataset import get_cache_filename
from cect.ConnectomeDataset import get_dataset_source_on_github
from cect import print_

from cect.Neurotransmitters import ACETYLCHOLINE
from cect.Neurotransmitters import DOPAMINE
from cect.Neurotransmitters import GABA
from cect.Neurotransmitters import GLUTAMATE
from cect.Neurotransmitters import OCTOPAMINE
from cect.Neurotransmitters import SEROTONIN
from cect.Neurotransmitters import TYRAMINE
from cect.Neurotransmitters import BETAINE
from cect.Neurotransmitters import ALL_KNOWN_CHEMICAL_NEUROTRANSMITTERS
from cect.Neurotransmitters import GENERIC_CHEM_SYN
from cect.Neurotransmitters import MONOAMINERGIC_SYN_CLASSES

from cect.Neurotransmitters import SEROTONIN_UPTAKE
from cect.Neurotransmitters import GABA_UPTAKE
from cect.Neurotransmitters import UNKNOWN_ORPHAN_NEUROTRANSMITTER
from cect.Neurotransmitters import UNKNOWN_MONOAMINERGIC_NEUROTRANSMITTER
from cect.Neurotransmitters import FIVE_HTP
from cect.Neurotransmitters import PEOH
from cect.Neurotransmitters import FIVE_HTP_FIVE_HT

from openpyxl import load_workbook

import os

spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"
filename = "%selife-95402-supp2-v1.xlsx" % spreadsheet_location


# BASIS_ANATOMICAL_CONN = "White_whole"
# BASIS_ANATOMICAL_CONN = "TestDataReader"
'''
BASIS_ANATOMICAL_CONN = ("Cook et al. 2019", "Cook2019HermReader")
BASIS_MONOAMINERGIC_CONN = ("Bentley et al. 2015", "WormNeuroAtlasMAReader")

READER_DESCRIPTION = (
    """A reader combining neurotransmitter atlas values from Wang et al. 2024 (source: %s) with basic anatomical connectivity information from %s, and monoaminergic receptor expression information from %s"""
    % (
        get_dataset_source_on_github(filename.split("/")[-1]),
        BASIS_ANATOMICAL_CONN[0],
        BASIS_MONOAMINERGIC_CONN[0],
    )
)'''
READER_DESCRIPTION = "????"


class Wang2024Reader(ConnectomeDataset):
    """
    Reader of data from multiple connectomes...
    """

    verbose = False

    spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"

    reader_description = "Not set..."

    dx1_set = False
    ef1_set = False
    dx3_set = False
    ef3_set = False

    def map_cell_name(self, cell_name: str) -> str:
        if cell_name == "DB1/3":
            return "DB1"
        elif cell_name == "DB3/1":
            return "DB3"
        elif cell_name == "DX1/2":
            if not self.dx1_set:
                self.dx1_set = True
                return "DX1"
            else:
                return "DX2"
        elif cell_name == "EF1/2":
            if not self.ef1_set:
                self.ef1_set = True
                return "EF1"
            else:
                return "EF2"
        elif cell_name == "DX3/4":
            if not self.dx3_set:
                self.dx3_set = True
                return "DX3"
            else:
                return "DX4"
        elif cell_name == "EF3/4":
            if not self.ef3_set:
                self.ef3_set = True
                return "EF3"
            else:
                return "EF4"
        elif len(cell_name) == 3 and cell_name[0:2] == "CA":
            return cell_name[0:2] + "0" + cell_name[2]
        elif len(cell_name) == 3 and cell_name[0:2] == "CP":
            return cell_name[0:2] + "0" + cell_name[2]
        else:
            if not is_any_neuron(cell_name):
                raise ValueError("Unknown cell name: %s" % cell_name)

            return cell_name

    def map_neurotransmitter(
        self, neurotransmitter: str, cat_1_present: bool, snf_3_present: bool
    ) -> str:
        """
        Maps neurotransmitter names to a standard format.
        """
        if neurotransmitter is None:
            return None
        neurotransmitter = neurotransmitter.strip()
        if (
            neurotransmitter == "ACh"
            or neurotransmitter == "ACh - NEW"
            or neurotransmitter == "*ACh"
            or neurotransmitter == "*ACh - NEW"
            or neurotransmitter == "ACh (new)"
            or neurotransmitter == "*ACh (new)"
        ):
            return ACETYLCHOLINE
        elif neurotransmitter == "DA":
            return DOPAMINE
        elif neurotransmitter == "GABA" or neurotransmitter == "*GABA":
            return GABA
        elif (
            neurotransmitter == "Glu"
            or neurotransmitter == "Glu - NEW"
            or neurotransmitter == "*Glu - NEW"
            or neurotransmitter == "*Glu (new)"
            or neurotransmitter == "Glu (new)"
            or neurotransmitter == "*Glu"
        ):
            return GLUTAMATE
        elif neurotransmitter == "octopamine":
            return OCTOPAMINE
        elif (
            neurotransmitter == "tyramine"
            or neurotransmitter == "tyramine (new)"
            or neurotransmitter == "tyramine (synthesis + uptake) - NEW"
        ):
            return TYRAMINE
        elif (
            neurotransmitter == "betaine (uptake)"
            or neurotransmitter == "betaine (uptake) - NEW"
            or neurotransmitter == "*betaine (uptake) - NEW"
            or neurotransmitter == "betaine (uptake) (new)"
            or neurotransmitter == "*betaine (uptake) (new)"
        ):
            if cat_1_present and snf_3_present:
                return BETAINE
            else:
                print_(
                    "  Note: Betaine neurotransmitter found without cat_1 or snf_3 present, returning None."
                )
                return None
        elif (
            neurotransmitter == "5-HT"
            or neurotransmitter == "5-HT (synthesis + uptake)"
            or neurotransmitter == "5-HT  (synthesis + uptake)"
            or neurotransmitter == "5-HT (new)"
            or neurotransmitter == "*5-HT (new)"
            or neurotransmitter == "5-HT (alternative synthesis/uptake mechanism)"
            or neurotransmitter
            == "male - 5-HT (alternative synthesis/uptake mechanism?)"
        ):
            return SEROTONIN
        elif neurotransmitter == "5-HTP - NEW" or neurotransmitter == "5-HTP (new)":
            return FIVE_HTP
        elif neurotransmitter == "5-HTP (synthesis) and 5-HT (uptake) (new)":
            return FIVE_HTP_FIVE_HT
        elif (
            neurotransmitter == "5-HT (uptake)" or neurotransmitter == "*5-HT (uptake)"
        ):
            return SEROTONIN_UPTAKE
        elif neurotransmitter == "GABA (uptake)":
            return GABA_UPTAKE
        elif neurotransmitter == "PEOH? (new)":
            return PEOH
        elif (
            neurotransmitter == "unknown (orphan)"
            or neurotransmitter.lower() == "unknown (orphan, unc-47 expression)"
            or neurotransmitter.lower() == "unknown (orphan; unc-47 positive)"
            or neurotransmitter == "unknown (orphan, unc-46 expression)"
        ):
            return UNKNOWN_ORPHAN_NEUROTRANSMITTER
        elif (
            neurotransmitter == "unknown monoamine? - NEW"
            or neurotransmitter == "unknown monoamine (new)"
            or neurotransmitter == "unknown monoamine? (new)"
            or neurotransmitter == "bas-1-depen unknown monoamine? - NEW"
            or neurotransmitter == "bas-1-depen unknown monoamine? (new)"
            or neurotransmitter == "*bas-1-depen unknown monoamine? (new)"
        ):
            return UNKNOWN_MONOAMINERGIC_NEUROTRANSMITTER
        else:
            raise ValueError("Unknown neurotransmitter: %s" % neurotransmitter)
            # return "NT_not_yet_supported__%s" % neurotransmitter.replace(' ', '_').replace('(', '_').replace(')', '_')  # neurotransmitter

    def __init__(self, sex, normalize_conn_numbers=True, include_monoamine_conns=True):
        ConnectomeDataset.__init__(self)

        sources = []
        self.all_neurotransmitters = {}

        if sex == "Hermaphrodite" or sex == "Male":
            sources.append(
                [
                    "%selife-95402-supp2-v1.xlsx" % self.spreadsheet_location,
                    "Supp File 2",
                    ("Cook et al. 2019 Hermaphrodite connectome", "Cook2019HermReader"),
                    ("Bentley et al. 2015", "WormNeuroAtlasMAReader"),
                ]
            )

        if sex == "Male":
            sources.append(
                [
                    "%selife-95402-supp3-v1.xlsx" % self.spreadsheet_location,
                    "Supp File 3",
                    ("Cook et al. 2019 Male connectome", "Cook2019MaleReader"),
                    ("Bentley et al. 2015", "WormNeuroAtlasMAReader"),
                ]
            )

        for source in sources:
            filename = source[0]
            sheet = source[1]
            BASIS_ANATOMICAL_CONN, BASIS_MONOAMINERGIC_CONN = source[2], source[3]

            wb = load_workbook(filename)
            print_("Opened the Excel file: " + filename)
            neurotransmitters = {}

            sheet = wb.get_sheet_by_name(sheet)
            rows = range(5, 307)
            col_offset = 0

            if "supp3" in filename:
                rows = range(6, 103)
                col_offset = 1
            for i in rows:
                print("Reading row %d" % i)

                cell_wang = sheet.cell(row=i, column=3).value
                if cell_wang is not None:
                    cell = self.map_cell_name(cell_wang)
                    cat_1_present = (
                        str(sheet.cell(row=i, column=9).fill.patternType) == "solid"
                    )
                    snf_3_present = "NEW" in str(sheet.cell(row=i, column=16).value)

                    nt_1 = self.map_neurotransmitter(
                        sheet.cell(row=i, column=21 + col_offset).value,
                        cat_1_present,
                        snf_3_present,
                    )
                    nt_2 = self.map_neurotransmitter(
                        sheet.cell(row=i, column=22 + col_offset).value,
                        cat_1_present,
                        snf_3_present,
                    )
                    nt_3 = self.map_neurotransmitter(
                        sheet.cell(row=i, column=23 + col_offset).value,
                        cat_1_present,
                        snf_3_present,
                    )

                    print_(
                        f"Reading row {i}: {cell}, NTs: {nt_1}, {nt_2}, {nt_3}; cat_1_present: {cat_1_present}, snf_3_present: {snf_3_present}"
                    )

                    nts = [nt_1]
                    if nt_2 is not None:
                        nts.append(nt_2)
                    if nt_3 is not None:
                        nts.append(nt_3)
                    print_("  - Cell: %s, nts: %s" % (cell, nts))
                    neurotransmitters[cell] = nts

            anatomical_conn_reader = load_connectome_dataset_file(
                get_cache_filename(BASIS_ANATOMICAL_CONN[1])
            )

            monoaminergic_conn_reader = load_connectome_dataset_file(
                get_cache_filename(BASIS_MONOAMINERGIC_CONN[1])
            )

            # neurons, muscles, other_cells, conns = self.read_all_data()

            anat_conns = anatomical_conn_reader.get_current_connection_info_list()

            print_("Adding %i conns from %s" % (len(anat_conns), BASIS_ANATOMICAL_CONN))
            for conn in anat_conns[:]:
                # print_("Original conn: %s" % conn)

                if is_any_neuron(conn.pre_cell) and conn.pre_cell in neurotransmitters:
                    if conn.synclass in ALL_KNOWN_CHEMICAL_NEUROTRANSMITTERS + [
                        GENERIC_CHEM_SYN
                    ]:
                        if normalize_conn_numbers:
                            conn.number = 1.0
                        for nt in neurotransmitters[conn.pre_cell]:
                            if nt in ALL_KNOWN_CHEMICAL_NEUROTRANSMITTERS:
                                conn.synclass = nt
                                # print_("    Adding new conn: %s" % conn)
                                self.add_connection_info(conn)
                    else:
                        print_(
                            "     Not a known chemical neurotransmitter: %s"
                            % conn.synclass
                        )
                else:
                    print_(
                        "     Not a neuron, or not in cells with known neurotransmitters..."
                    )

            if include_monoamine_conns:
                monoamine_conns = (
                    monoaminergic_conn_reader.get_current_connection_info_list()
                )

                print_(
                    "Adding %i conns from %s"
                    % (len(monoamine_conns), BASIS_MONOAMINERGIC_CONN)
                )
                for conn in monoamine_conns[:]:
                    print_("Original conn: %s" % conn)

                    if (
                        is_any_neuron(conn.pre_cell)
                        and conn.pre_cell in neurotransmitters
                    ):
                        if normalize_conn_numbers:
                            conn.number = 1.0
                        for nt in neurotransmitters[conn.pre_cell]:
                            if nt in MONOAMINERGIC_SYN_CLASSES:
                                conn.synclass = nt
                                print_("    Adding new conn: %s" % conn)
                                self.add_connection_info(conn)

                    else:
                        print_(
                            "     Not a neuron, or not in cells with known neurotransmitters..."
                        )

            self.all_neurotransmitters.update(neurotransmitters)

        self.reader_description = (
            """A reader combining neurotransmitter atlas values from Wang et al. 2024 (source: %s) with basic anatomical connectivity information from %s, and monoaminergic receptor expression information from %s"""
            % (
                "; ".join(
                    [get_dataset_source_on_github(f[0].split("/")[-1]) for f in sources]
                ),
                BASIS_ANATOMICAL_CONN[0],
                BASIS_MONOAMINERGIC_CONN[0],
            )
        )

    def read_data(self):
        return self._read_data()

    def read_muscle_data(self):
        return self._read_muscle_data()


def get_instance(from_cache=LOAD_READERS_FROM_CACHE_BY_DEFAULT):
    if from_cache:
        from cect.ConnectomeDataset import (
            load_connectome_dataset_file,
            get_cache_filename,
        )

        return load_connectome_dataset_file(get_cache_filename(__name__.split(".")[1]))
    else:
        global READER_DESCRIPTION
        inst = Wang2024Reader(sex="Hermaphrodite")
        READER_DESCRIPTION = inst.reader_description
        return inst


def main():
    tdr_instance = get_instance(from_cache=False)

    print(tdr_instance.summary(list_pre_cells=False))
    print(tdr_instance.reader_description)
    quit()

    print(tdr_instance.all_neurotransmitters)

    cells, neuron_conns = tdr_instance.read_data()

    neurons2muscles, muscles, muscle_conns = tdr_instance.read_muscle_data()

    analyse_connections(
        cells,
        neuron_conns,
        neurons2muscles,
        muscles,
        muscle_conns,
        print_details_on=["ADFL"],
    )

    print_(" -- Finished analysing connections using: %s" % os.path.basename(__file__))

    """
    
    from cect.ConnectomeView import COOK_FIG3_VIEW as 
    from cect.ConnectomeView import LOCOMOTION_3_VIEW as view
    from cect.ConnectomeView import LOCOMOTION_1_VIEW as view
    """
    # from cect.ConnectomeView import SOCIAL_VIEW as view
    # from cect.ConnectomeView import LOCOMOTION_2_VIEW as view
    from cect.ConnectomeView import RAW_VIEW as view
    # from cect.ConnectomeView import NEURONS_VIEW as view

    cds2 = tdr_instance.get_connectome_view(view)

    print(cds2.summary(list_pre_cells=True))

    # fig = cds2.to_plotly_hive_plot_fig(list(view.synclass_sets.keys())[0], view)

    """
    fig = cds2.to_plotly_graph_fig(list(view.synclass_sets.keys())[0], view)
    """

    from cect.Cells import GLUTAMATE

    fig, _ = cds2.to_plotly_matrix_fig(
        GLUTAMATE,
        view,
    )
    import plotly.io as pio

    pio.renderers.default = "browser"
    import sys

    if "-nogui" not in sys.argv:
        fig.show()


if __name__ == "__main__":
    main()
