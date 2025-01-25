from cect.ConnectomeReader import ConnectionInfo
from cect.ConnectomeReader import analyse_connections
from cect.Cells import convert_to_preferred_muscle_name
from cect.Cells import is_herm_neuron
from cect.Cells import is_potential_muscle
from cect.Cells import is_known_muscle
from cect.Cells import remove_leading_index_zero

from cect.ConnectomeDataset import ConnectomeDataset

from cect.Cells import GENERIC_CHEM_SYN
from cect.Cells import GENERIC_ELEC_SYN

from openpyxl import load_workbook
import os
from cect import print_

spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"


def fix_witvliet_cell_naming(cell):
    if cell == "excgl":
        return "exc_gl"
    else:
        return cell


class WitvlietDataReader(ConnectomeDataset):
    """Reader for datasets from [Witvliet et al. 2021](../../Witvliet_2021.md)"""

    verbose = False

    def __init__(self, spreadsheet):
        ConnectomeDataset.__init__(self)
        self.filename = "%s%s" % (spreadsheet_location, spreadsheet)

        neurons, muscles, other_cells, conns = self.read_all_data()

        for conn in conns:
            self.add_connection_info(conn)

    def read_all_data(self):
        neurons = set([])
        muscles = set([])
        other_cells = set([])
        conns = []

        wb = load_workbook(self.filename)
        sheet = wb.worksheets[0]
        print_("Opened the Excel file: " + self.filename)

        for row in sheet.iter_rows(
            min_row=2, values_only=True
        ):  # Assuming data starts from the second row
            pre = str(row[0])
            post = str(row[1])

            pre = fix_witvliet_cell_naming(remove_leading_index_zero(pre))
            post = fix_witvliet_cell_naming(remove_leading_index_zero(post))

            if is_potential_muscle(pre):
                pre = convert_to_preferred_muscle_name(pre)

            if is_potential_muscle(post):
                post = convert_to_preferred_muscle_name(post)

            syntype = str(row[2])
            num = int(row[3])

            if self.verbose and num > 0:
                print("Conn %s -> %s #%i" % (pre, post, num))

            synclass = GENERIC_ELEC_SYN if "electrical" in syntype else GENERIC_CHEM_SYN
            if synclass == GENERIC_ELEC_SYN:
                conns.append(ConnectionInfo(post, pre, num, syntype, synclass))

            conns.append(ConnectionInfo(pre, post, num, syntype, synclass))

            for p in [pre, post]:
                if is_herm_neuron(p):
                    neurons.add(pre)
                elif is_known_muscle(p):
                    muscles.add(pre)
                else:
                    other_cells.add(p)

        return list(neurons), list(muscles), list(other_cells), conns


def main1():
    wdr = WitvlietDataReader("witvliet_2020_7.xlsx")

    cells, neuron_conns = wdr.read_data()
    neurons2muscles, muscles, muscle_conns = wdr.read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)


def main2():
    wdr = WitvlietDataReader("witvliet_2020_8.xlsx")

    cells, neuron_conns = wdr.read_data()
    neurons2muscles, muscles, muscle_conns = wdr.read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)


if __name__ == "__main__":
    main1()
    main2()
