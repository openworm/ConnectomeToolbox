# -*- coding: utf-8 -*-

############################################################

#    A script to read the values of Cook et al 2019

############################################################


from cect.ConnectomeReader import ConnectionInfo
from cect.ConnectomeReader import analyse_connections
from cect.Cells import convert_to_preferred_muscle_name
from cect.Cells import is_any_neuron
from cect.Cells import remove_leading_index_zero
from cect.Cells import is_potential_muscle
from cect.Cells import is_known_muscle

from cect.ConnectomeDataset import ConnectomeDataset

from cect.Cells import GENERIC_CHEM_SYN
from cect.Cells import GENERIC_ELEC_SYN

from openpyxl import load_workbook

import os
import numpy as np

from cect import print_

HERM_CHEM = "hermaphrodite chemical"
HERM_GAP_SYMM = "herm gap jn symmetric"
MALE_CHEM = "male chemical"
MALE_GAP_SYMM = "male gap jn symmetric"

SEX_SPECIFIC_SHEETS = {
    "Hermaphodite": [HERM_CHEM, HERM_GAP_SYMM],
    "Male": [MALE_CHEM, MALE_GAP_SYMM],
}

pre_range = {
    HERM_CHEM: range(4, 304),
    HERM_GAP_SYMM: range(4, 472),
    MALE_CHEM: range(4, 386),
    MALE_GAP_SYMM: range(4, 472),
}
post_range = {
    HERM_CHEM: range(4, 457),
    HERM_GAP_SYMM: range(4, 472),
    MALE_CHEM: range(4, 579),
    MALE_GAP_SYMM: range(4, 589),
}


def get_synclass(cell, syntype):
    if syntype == "GapJunction":
        return GENERIC_ELEC_SYN
    else:
        '''# dirty hack
        if cell.startswith("DD") or cell.startswith("VD"):
            return "GABA"'''
        return GENERIC_CHEM_SYN


class Cook2019DataReader(ConnectomeDataset):
    """
    Reader of data from Cook et al. 2019 - Whole-animal connectomes of both Caeonrhabditis elegans sexes
    """

    spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"
    filename = "%sSI 5 Connectome adjacency matrices.xlsx" % spreadsheet_location

    verbose = False

    def __init__(self, sex):
        ConnectomeDataset.__init__(self)
        self.sex = sex

        wb = load_workbook(self.filename)
        print_("Opened the Excel file: " + self.filename)

        self.pre_cells = {}
        self.post_cells = {}
        self.conn_nums = {}

        for conn_type in SEX_SPECIFIC_SHEETS[self.sex]:
            sheet = wb.get_sheet_by_name(conn_type)
            print_("Looking at sheet: %s" % conn_type)

            self.pre_cells[conn_type] = []
            self.post_cells[conn_type] = []

            for i in pre_range[conn_type]:
                self.pre_cells[conn_type].append(sheet["C%i" % i].value)

            if self.verbose:
                print_(
                    " - Pre cells for %s (%i):\n%s"
                    % (
                        conn_type,
                        len(self.pre_cells[conn_type]),
                        self.pre_cells[conn_type],
                    )
                )

            for i in post_range[conn_type]:
                self.post_cells[conn_type].append(sheet.cell(row=3, column=i).value)

            if self.verbose:
                print_(
                    " - Post cells for %s (%i):\n%s"
                    % (
                        conn_type,
                        len(self.post_cells[conn_type]),
                        self.post_cells[conn_type],
                    )
                )

            self.conn_nums[conn_type] = np.zeros(
                [len(self.pre_cells[conn_type]), len(self.post_cells[conn_type])],
                dtype=int,
            )

            for i in range(len(self.pre_cells[conn_type])):
                for j in range(len(self.post_cells[conn_type])):
                    row = 4 + i
                    col = 4 + j
                    val = sheet.cell(row=row, column=col).value
                    # print("Cell (%i,%i) [row %i, col %i] = %s" % (i, j, row, col, val))
                    if val is not None:
                        self.conn_nums[conn_type][i, j] = int(val)

            if self.verbose:
                print_(
                    " - Conns for %s (%s):\n%s"
                    % (
                        conn_type,
                        self.conn_nums[conn_type].shape,
                        self.conn_nums[conn_type],
                    )
                )

        neurons, muscles, other_cells, conns = self.read_all_data()

        for conn in conns:
            self.add_connection_info(conn)

    def read_all_data(self):
        """
        Returns:
            Tuple[list, list, list, list]: List of neurons, muscles), other cells and connections which have been read in
        """

        neurons = set([])
        muscles = set([])
        other_cells = set([])
        conns = []

        for conn_type in SEX_SPECIFIC_SHEETS[self.sex]:
            for pre_index in range(len(self.pre_cells[conn_type])):
                for post_index in range(len(self.post_cells[conn_type])):
                    num = self.conn_nums[conn_type][pre_index, post_index]

                    pre = remove_leading_index_zero(
                        self.pre_cells[conn_type][pre_index]
                    )
                    post = remove_leading_index_zero(
                        self.post_cells[conn_type][post_index]
                    )
                    if self.verbose and num > 0:
                        print_("Conn %s -> %s #%i" % (pre, post, num))

                    if is_potential_muscle(pre):
                        pre = convert_to_preferred_muscle_name(pre)

                    if is_potential_muscle(post):
                        post = convert_to_preferred_muscle_name(post)

                    if num > 0:
                        syntype = "Send" if "chemical" in conn_type else "GapJunction"
                        synclass = get_synclass(pre, syntype)

                        ci = ConnectionInfo(pre, post, num, syntype, synclass)
                        if self.verbose:
                            print_("Conn: %s" % (ci))
                        conns.append(ci)

                        for p in [pre, post]:
                            if is_any_neuron(p):
                                neurons.add(pre)
                            elif is_known_muscle(p):
                                muscles.add(pre)
                            else:
                                other_cells.add(p)

        return list(neurons), list(muscles), list(other_cells), conns


def main():
    tdr_instance = Cook2019DataReader("Hermaphodite")

    cells, neuron_conns = tdr_instance.read_data()

    neurons2muscles, muscles, muscle_conns = tdr_instance.read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)

    print_(" -- Finished analysing connections using: %s" % os.path.basename(__file__))

    print(tdr_instance.summary())


if __name__ == "__main__":
    main()
