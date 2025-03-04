from cect.ConnectomeReader import ConnectionInfo
from cect.ConnectomeReader import analyse_connections
from cect.ConnectomeDataset import ConnectomeDataset
from cect.ConnectomeDataset import get_dataset_source_on_github
from cect.ConnectomeDataset import LOAD_READERS_FROM_CACHE_BY_DEFAULT

from cect.Cells import GENERIC_CHEM_SYN
from cect.Cells import GENERIC_ELEC_SYN

from openpyxl import load_workbook

import os
import sys

from cect import print_

spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"
spreadsheet_name = "NeuronConnect.xlsx"  # has old name...
spreadsheet_name = "NeuronConnectFormatted.xlsx"
filename = "%s%s" % (spreadsheet_location, spreadsheet_name)

READER_DESCRIPTION = (
    """Data extracted from %s for neuronal connectivity"""
    % get_dataset_source_on_github(filename.split("/")[-1])
)

NMJ_ENDPOINT = "NMJ"


class VarshneyDataReader(ConnectomeDataset):
    """Reader for Varshney et al. 2011 connectivity dataset"""

    def __init__(self):
        ConnectomeDataset.__init__(self)

        cells, neuron_conns = self.read_data()
        for conn in neuron_conns:
            self.add_connection_info(conn, append_existing_connections=True)

    def read_data(self):
        cells = []
        conns = []

        wb = load_workbook(filename)
        sheet = wb.worksheets[0]
        print_("Opened the Excel file: " + filename)

        for row in sheet.iter_rows(
            min_row=2, values_only=True
        ):  # Assuming data starts from the second row
            pre = str(row[0])
            post = str(row[1])

            if not post == NMJ_ENDPOINT:
                syntype = str(row[2])
                num = int(row[3])
                synclass = (
                    GENERIC_ELEC_SYN
                    if syntype == "EJ"
                    else GENERIC_CHEM_SYN
                    if (syntype == "Sp" or syntype == "S")
                    else None
                )

                if synclass is not None:
                    conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
                    if pre not in cells:
                        cells.append(pre)
                    if post not in cells:
                        cells.append(post)

        return cells, conns

    def read_muscle_data(self):
        conns = []
        neurons = []
        muscles = []

        return neurons, muscles, conns


def get_instance(from_cache=LOAD_READERS_FROM_CACHE_BY_DEFAULT):
    if from_cache:
        from cect.ConnectomeDataset import (
            load_connectome_dataset_file,
            get_cache_filename,
        )

        return load_connectome_dataset_file(get_cache_filename(__name__.split(".")[1]))
    else:
        return VarshneyDataReader()


"""
read_data = my_instance.read_data
read_muscle_data = my_instance.read_muscle_data"""


def main():
    my_instance = get_instance()

    cells, neuron_conns = my_instance.read_data()
    neurons2muscles, muscles, muscle_conns = my_instance.read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)

    print_(" -- Finished analysing connections using: %s" % os.path.basename(__file__))

    if "-nogui" not in sys.argv:
        my_instance.connection_number_plot(GENERIC_CHEM_SYN)


if __name__ == "__main__":
    main()
