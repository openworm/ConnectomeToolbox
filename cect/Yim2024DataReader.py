# -*- coding: utf-8 -*-

############################################################

#    A script to read the values of Yim et al. 2024

############################################################


from cect.ConnectomeReader import ConnectionInfo
from cect.ConnectomeReader import analyse_connections
from cect.Cells import convert_to_preferred_muscle_name
from cect.Cells import is_any_neuron
from cect.Cells import remove_leading_index_zero
from cect.Cells import is_potential_muscle
from cect.Cells import is_known_muscle
from cect.ConnectomeDataset import get_dataset_source_on_github
from cect.ConnectomeDataset import LOAD_READERS_FROM_CACHE_BY_DEFAULT

from cect.ConnectomeDataset import ConnectomeDataset

from cect.Neurotransmitters import CONTACTOME_SYN_TYPE
from cect.Neurotransmitters import CONTACTOME_SYN_CLASS

# ruff: noqa: F401
from cect.Neurotransmitters import GENERIC_CHEM_SYN
from cect.Neurotransmitters import GENERIC_ELEC_SYN

from openpyxl import load_workbook

import os
import numpy as np

from cect import print_

DAUER_NON_NORM = "Dauer"
DAUER_NORM = "Dauer_normalized"


NAME = "Yim2024"

pre_range = range(3, 225)
post_range = range(3, 225)


def get_synclass(cell, normalized):
    if normalized:
        return GENERIC_CHEM_SYN
    else:
        return CONTACTOME_SYN_CLASS


spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"
filename = "%s41467_2024_45943_MOESM6_ESM.xlsx" % spreadsheet_location


READER_DESCRIPTION = (
    """Data extracted from %s Yim et al. 2024 on Dauer connectome (Normalized)"""
    % get_dataset_source_on_github(filename.split("/")[-1])
)


class Yim2024DataReader(ConnectomeDataset):
    """
    Reader of data from Yim et al. 2024 - Dauer connectome
    """

    verbose = False

    def __init__(self, normalized=True):
        ConnectomeDataset.__init__(self)

        conn_type = DAUER_NORM if normalized else DAUER_NON_NORM

        print_(f"Opening sheet {conn_type} in the Excel file: {filename}")

        wb = load_workbook(filename)

        self.pre_cells = {}
        self.post_cells = {}
        self.conn_nums = {}
        self.normalized = normalized

        sheet = wb.get_sheet_by_name(conn_type)

        self.pre_cells[conn_type] = []
        self.post_cells[conn_type] = []

        for i in pre_range:
            self.pre_cells[conn_type].append(sheet["A%i" % i].value)

        if self.verbose:
            print_(
                " - Pre cells for %s (%i):\n%s"
                % (
                    conn_type,
                    len(self.pre_cells[conn_type]),
                    self.pre_cells[conn_type],
                )
            )

        for i in post_range:
            self.post_cells[conn_type].append(sheet.cell(row=1, column=i).value)

        if self.verbose:
            print_(
                " - Post cells for %s (%i):\n%s"
                % (
                    conn_type,
                    len(self.post_cells[conn_type]),
                    self.post_cells[conn_type],
                )
            )

        self.conn_nums[conn_type] = np.zeros(
            [len(self.pre_cells[conn_type]), len(self.post_cells[conn_type])],
            dtype=float,
        )

        for i in range(len(self.pre_cells[conn_type])):
            for j in range(len(self.post_cells[conn_type])):
                row = 3 + i
                col = 3 + j
                val = sheet.cell(row=row, column=col).value
                if val != 0:
                    print_("Cell (%i,%i) [row %i, col %i] = %s" % (i, j, row, col, val))
                if val is not None:
                    self.conn_nums[conn_type][i, j] = val

        if self.verbose:
            print_(
                " - Conns for %s (%s):\n%s"
                % (
                    conn_type,
                    self.conn_nums[conn_type].shape,
                    self.conn_nums[conn_type],
                )
            )

        neurons, muscles, other_cells, conns = self.read_all_data()

        for conn in conns:
            self.add_connection_info(conn)

    def read_data(self):
        return self._read_data()

    def read_muscle_data(self):
        return self._read_muscle_data()

    def read_all_data(self):
        """
        Returns:
            Tuple[list, list, list, list]: List of neurons, muscles), other cells and connections which have been read in
        """

        neurons = set([])
        muscles = set([])
        other_cells = set([])
        conns = []

        conn_type = DAUER_NORM if self.normalized else DAUER_NON_NORM

        for pre_index in range(len(self.pre_cells[conn_type])):
            for post_index in range(len(self.post_cells[conn_type])):
                num = self.conn_nums[conn_type][pre_index, post_index]

                pre = remove_leading_index_zero(self.pre_cells[conn_type][pre_index])
                post = remove_leading_index_zero(self.post_cells[conn_type][post_index])
                if self.verbose and num > 0:
                    print_("Conn %s -> %s #%i" % (pre, post, num))

                if is_potential_muscle(pre):
                    pre = convert_to_preferred_muscle_name(pre)

                if is_potential_muscle(post):
                    post = convert_to_preferred_muscle_name(post)

                if num > 0:
                    syntype = CONTACTOME_SYN_TYPE
                    synclass = get_synclass(pre, self.normalized)

                    ci = ConnectionInfo(pre, post, num, syntype, synclass)
                    if self.verbose:
                        print_("Conn: %s" % (ci))
                    conns.append(ci)

                    for p in [pre, post]:
                        if is_any_neuron(p):
                            neurons.add(pre)
                        elif is_known_muscle(p):
                            muscles.add(pre)
                        else:
                            other_cells.add(p)

        return list(neurons), list(muscles), list(other_cells), conns


def get_instance(from_cache=LOAD_READERS_FROM_CACHE_BY_DEFAULT):
    """Uses ``Yim2024DataReader`` to load data on dauer connectome

    Returns:
        Yim2024DataReader: The initialized connectome reader
    """
    if from_cache:
        from cect.ConnectomeDataset import (
            load_connectome_dataset_file,
            get_cache_filename,
        )

        return load_connectome_dataset_file(
            get_cache_filename(__file__.split("/")[-1].split(".")[0])
        )
    else:
        return Yim2024DataReader(normalized=True)


def main():
    tdr_instance = get_instance(from_cache=False)

    # analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)

    print_(" -- Finished analysing connections using: %s" % os.path.basename(__file__))

    print(tdr_instance.summary())

    from cect.ConnectomeView import RAW_VIEW as view
    # from cect.ConnectomeView import PHARYNX_VIEW as view
    # from cect.ConnectomeView import NEURONS_VIEW as view

    print("=======================")
    cds2 = tdr_instance.get_connectome_view(view)
    print(cds2.summary(list_pre_cells=False))

    print("Plotting view: %s" % view)
    fig, _ = cds2.to_plotly_matrix_fig(
        "Chemical",
        view,
    )
    import plotly.io as pio

    pio.renderers.default = "browser"
    import sys

    if "-nogui" not in sys.argv:
        fig.show()


if __name__ == "__main__":
    main()
